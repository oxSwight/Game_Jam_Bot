"""Command-level tests: drive the aiogram handlers with a fake Message that
records what the bot would send, exercising arg parsing + happy/error paths."""

from types import SimpleNamespace

from app.handlers import admin_extra
from app.models.application import ApplicationStatus
from tests.conftest import make_payload


class FakeMessage:
    def __init__(self, text: str = "", user_id: int = 111):
        self.text = text
        self.html_text = text
        self.from_user = SimpleNamespace(id=user_id, username="admin", language_code="ru")
        self.chat = SimpleNamespace(id=user_id)
        self.answers: list[str] = []
        self.documents: list[tuple] = []

    async def answer(self, text, **kw):
        self.answers.append(text)

    async def answer_document(self, document, caption=None, **kw):
        self.documents.append((document, caption))


async def _submit_approved(services, session, n):
    for i in range(n):
        read = await services.applications.submit_registration(
            make_payload(telegram_id=500 + i, nickname=f"Play{i}", email=f"h{i}@x.com")
        )
        await services.applications.update_status(read.id, ApplicationStatus.APPROVED)
    await session.commit()


# ------------------------------ /stats ------------------------------ #
async def test_cmd_stats_reports_counts(services, session):
    await _submit_approved(services, session, 2)
    msg = FakeMessage("/stats")
    await admin_extra.cmd_stats(msg, services)
    assert msg.answers
    assert "Статистика" in msg.answers[0]
    assert "Одобрено" in msg.answers[0]


# ------------------------------ /export ------------------------------ #
async def test_cmd_export_empty(services):
    msg = FakeMessage("/export")
    await admin_extra.cmd_export(msg, services)
    assert "нечего" in msg.answers[0].lower()


async def test_cmd_export_produces_xlsx_and_csv(services, session):
    await _submit_approved(services, session, 1)
    msg = FakeMessage("/export")
    await admin_extra.cmd_export(msg, services)
    assert [doc.filename for doc, _ in msg.documents] == [
        "applications.xlsx",
        "applications.csv",
    ]
    assert "1" in msg.documents[0][1]


async def test_cmd_export_neutralises_formula_injection(services, session):
    # A nickname that is a spreadsheet formula must not stay live in either file.
    await services.applications.submit_registration(
        make_payload(telegram_id=700, nickname="=1+1", email="evil@x.com")
    )
    await session.commit()

    msg = FakeMessage("/export")
    await admin_extra.cmd_export(msg, services)

    xlsx_rows = _read_xlsx_rows(msg.documents[0][0].data)
    nickname_col = xlsx_rows[0].index("nickname")
    assert xlsx_rows[1][nickname_col] == "'=1+1"   # neutralised: inert text

    csv_text = msg.documents[1][0].data.decode("utf-8")
    assert "'=1+1" in csv_text            # neutralised: leading quote → inert text
    assert ";=1+1" not in csv_text        # the bare formula must not appear as a cell


def _read_xlsx_rows(data: bytes) -> list[list]:
    from io import BytesIO

    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(data), read_only=True).active
    return [["" if v is None else v for v in row] for row in ws.iter_rows(values_only=True)]


async def test_cmd_export_multiselects_stay_in_their_columns(services, session):
    """Regression for the export column drift: multi-selects are joined with
    '; ' - under the old comma-delimited CSV those cells were unquoted, and a
    ';'-splitting importer (Russian-locale Excel) tore each row into a different
    number of columns. Round-trip both files and assert the grid stays aligned."""
    import csv as _csv
    import io as _io

    from app.schemas.registration import RegistrationCreate

    payload = RegistrationCreate.from_fsm_data(
        data={
            "nickname": "Multi",
            "email": "multi@x.com",
            "category_id": "programming",
            "category_title": "Programming / Engineering",
            "roles": ["gameplay_programmer", "general_programmer"],
            "experience_level": "beginner",
            "engine": ["Unity", "Godot"],
            "engine_other": None,
            "tools": ["Blender", "Photoshop"],
            "tools_other": None,
            "motivations": ["Learning", "Portfolio", "Team experience"],
            "strengths": ["Прототипировать", "Балансировать"],
            "consent": True,
        },
        telegram_id=800,
        telegram_username="multi",
    )
    await services.applications.submit_registration(payload)
    await session.commit()

    msg = FakeMessage("/export")
    await admin_extra.cmd_export(msg, services)

    xlsx_rows = _read_xlsx_rows(msg.documents[0][0].data)
    csv_rows = list(
        _csv.reader(
            _io.StringIO(msg.documents[1][0].data.decode("utf-8-sig")), delimiter=";"
        )
    )

    for rows in (xlsx_rows, csv_rows):
        header, data_row = rows[0], rows[1]
        assert len(rows) == 2
        # Every parsed row has exactly the header's column count - no drift.
        assert len(data_row) == len(header) == len(admin_extra._EXPORT_HEADER)
        assert data_row[header.index("nickname")] == "Multi"
        assert data_row[header.index("motivations")] == "Learning; Portfolio; Team experience"
        assert data_row[header.index("engine")] == "Unity, Godot"
        assert data_row[header.index("tools")] == "Blender, Photoshop"
        assert data_row[header.index("strengths")] == "Прототипировать; Балансировать"
        assert data_row[header.index("created_at")]  # last column landed in place
